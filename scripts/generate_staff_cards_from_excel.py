"""
Generate a combined PDF of HERO digital membership cards from an Excel sheet.

Usage:
    python scripts/generate_staff_cards_from_excel.py
"""

from __future__ import annotations

import json
from io import BytesIO
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

import qrcode


EXCEL_PATH = Path(
    r"C:\Users\Matloob-Alam\Downloads\Staff Registration - Hero - June 2026 (003) (005).xlsx"
)
OUTPUT_PATH = Path.home() / "Downloads" / "HERO_Staff_Digital_Membership_Cards.pdf"
LOGO_PATH = Path(__file__).resolve().parent.parent / "public" / "brand" / "hero-logo.png"
COMPANY_NAME = "HERO Client Rescue S.A."


def to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def make_qr_data_uri(membership_id: str) -> ImageReader:
    payload = json.dumps({"v": 1, "membership": membership_id, "token": ""}, separators=(",", ":"))
    qr = qrcode.QRCode(version=2, box_size=8, border=2)
    qr.add_data(payload)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    pil_img = img.get_image() if hasattr(img, "get_image") else img
    buffer = BytesIO()
    pil_img.save(buffer, format="PNG")
    buffer.seek(0)
    return ImageReader(buffer)


def load_rows(xlsx_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records: list[dict[str, str]] = []
    for row_idx in range(3, ws.max_row + 1):
        name = ws.cell(row_idx, 3).value
        plan = ws.cell(row_idx, 6).value
        expiration_raw = ws.cell(row_idx, 7).value
        membership_id = ws.cell(row_idx, 8).value

        if not name or not membership_id:
            continue

        expiration = to_date(expiration_raw)
        expiration_label = expiration.strftime("%B %d, %Y") if expiration else "—"
        status = "Active" if (expiration and expiration >= date.today()) else "Expired"

        records.append(
            {
                "name": str(name).strip(),
                "plan": str(plan).strip() if plan else "—",
                "membership_id": str(membership_id).strip(),
                "expiration": expiration_label,
                "status": status,
            }
        )

    return records


def draw_card(c: canvas.Canvas, row: dict[str, str]) -> None:
    page_w, page_h = A4

    card_w = 170 * mm
    card_h = 110 * mm
    x = (page_w - card_w) / 2
    y = (page_h - card_h) / 2

    c.setFillColor(colors.HexColor("#e4e4e8"))
    c.roundRect(x, y, card_w, card_h, 14, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor("#c9ced6"))
    c.roundRect(x, y, card_w, card_h, 14, fill=0, stroke=1)

    top_y = y + card_h - 18 * mm
    if LOGO_PATH.is_file():
        c.drawImage(str(LOGO_PATH), x + 8 * mm, top_y - 2 * mm, width=14 * mm, height=12 * mm, preserveAspectRatio=True, mask="auto")
    else:
        c.setFillColor(colors.HexColor("#00838f"))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x + 8 * mm, top_y + 2 * mm, "HERO")
    c.setFillColor(colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x + 25 * mm, top_y + 4 * mm, COMPANY_NAME)

    c.setFillColor(colors.HexColor("#334155"))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(x + card_w - 30 * mm, top_y + 7 * mm, "STATUS")
    c.setFillColor(colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(x + card_w - 8 * mm, top_y + 3 * mm, row["status"])

    c.setFillColor(colors.HexColor("#334155"))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(x + 8 * mm, y + card_h - 31 * mm, "MEMBER'S NAME")
    c.setFillColor(colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 18)
    c.drawString(x + 8 * mm, y + card_h - 39 * mm, row["name"][:36])

    c.setFillColor(colors.HexColor("#334155"))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(x + 8 * mm, y + card_h - 50 * mm, "MEMBERSHIP TYPE")
    c.setFillColor(colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x + 8 * mm, y + card_h - 56 * mm, row["plan"][:42])

    c.setFillColor(colors.HexColor("#334155"))
    c.setFont("Helvetica-Bold", 7)
    c.drawRightString(x + card_w - 8 * mm, y + card_h - 50 * mm, "MEMBERSHIP ID")
    c.setFillColor(colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(x + card_w - 8 * mm, y + card_h - 56 * mm, row["membership_id"])

    c.setFillColor(colors.HexColor("#334155"))
    c.setFont("Helvetica-Bold", 7)
    c.drawString(x + 8 * mm, y + card_h - 66 * mm, "EXPIRATION")
    c.setFillColor(colors.HexColor("#0f172a"))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x + 8 * mm, y + card_h - 72 * mm, row["expiration"])

    qr = make_qr_data_uri(row["membership_id"])
    qr_size = 42 * mm
    qr_x = x + (card_w - qr_size) / 2
    qr_y = y + 8 * mm
    c.setFillColor(colors.white)
    c.roundRect(qr_x - 3, qr_y - 3, qr_size + 6, qr_size + 6, 6, fill=1, stroke=0)
    c.setStrokeColor(colors.HexColor("#cbd5e1"))
    c.roundRect(qr_x - 3, qr_y - 3, qr_size + 6, qr_size + 6, 6, fill=0, stroke=1)
    c.drawImage(qr, qr_x, qr_y, width=qr_size, height=qr_size, preserveAspectRatio=True, mask="auto")


def main() -> int:
    if not EXCEL_PATH.is_file():
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    rows = load_rows(EXCEL_PATH)
    if not rows:
        raise RuntimeError("No valid rows found. Check Name and Membership ID columns.")

    c = canvas.Canvas(str(OUTPUT_PATH), pagesize=A4)
    for idx, row in enumerate(rows):
        draw_card(c, row)
        if idx < len(rows) - 1:
            c.showPage()
    c.save()

    print(f"Generated {len(rows)} membership cards.")
    print(f"Saved: {OUTPUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
